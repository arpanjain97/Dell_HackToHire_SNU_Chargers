import os
import openpyxl 

wb = openpyxl.load_workbook('dell_dataset_final.xlsx',data_only=True)
ws = wb["dell_dataset_final"]

wb1 =  openpyxl.load_workbook('dtv.xlsx',data_only=True)
ws1 = wb1["Sheet1"]

for i in range(2,4437):

    year = str(ws["G" + str(i)].value)
    ram = str(ws["M" + str(i)].value)
    pro = str(ws["N" + str(i)].value)
    ssd = str(ws["Q" + str(i)].value)
    ps = 0
    rs = 0
    ss = 0
    if(year=="2018"):
        if(ram>='32'):
            rs = 3
        elif(ram>='16'):
            rs = 2
        elif(ram>='8'):
            rs = 1
        if(ssd>='1024'):
            ss = 3
        elif(ssd>='512'):
            ss = 2
        else:
            ss = 1
        if(pro=='i9'):
            ps = 3
        elif(pro=='i7'):
            ps = 2
        else:
            ps = 1
    elif(year=='2017' or year=='2016'):
        if(ram>='16'):
            rs = 3
        elif(ram>='8'):
            rs = 2
        elif(ram>='4'):
            rs = 1
        if(ssd>='512'):
            ss = 3
        elif(ssd>='256'):
            ss = 2
        else:
            ss = 1
        if(pro=='i7'):
            ps = 3
        elif(pro=='i5'):
            ps = 2
        else:
            ps = 1

    avg = (ps+ss+rs)/3
    T = ws1['B' +str(i)]
    cid =  str(ws["A"+str(i)].value)
    Cid = ws1['A' + str(i)]
    Cid.value = cid
    if(avg>=3):
   
        T.value = "3"
    elif(avg>=2):
        T.value = "2"
         
    else:
        T.value = "1"
           
wb1.save('dtv.xlsx')







