import os
import openpyxl 

wb = openpyxl.load_workbook('rfm_set.xlsx',data_only=True)
ws = wb["rfm_set"]
ws["J1"] = "Recommendations"
wb1 = openpyxl.load_workbook('dell_dataset_final.xlsx',data_only=True)
ws1 = wb1['dell_dataset_final']

for i in range(2,4437):

    recent = str(ws[("E" + str(i))].value)
    count =  str(ws[("F" + str(i))].value)
    spent =  str(ws[("G" + str(i))].value)
    warranty = str(ws1['S' + str(i)].value)

    recommend = ws['J' + str(i)]
    
    if(recent=="1" and count=="1" and spent=="1"):
        recommend.value = "Accessories_OtherProducts_FestiveOffers"
    elif(count<="2"):
        if(recent>="2"):
            if(spent<="3"):
                 recommend.value = "NewExpensiveProducts_Accessories"
            else:
              recommend.value = "NewProducts_FestivalOffers_Accessories"
        else:
              recommend.value  = "Accessories"
    elif(spent<="2"):
        if(recent>="2"):
             recommend.value  = "NewExpensiveProducts_Accessories_FestivalOffers"
        else:
             recommend.value  = "NewProducts_Accessories_FestivalOffers"
    elif(recent>"2"):
        if(spent=="1" and count=="1"):
             recommend.value  = "NewExpensive_products_Accesories_Discount_FestivalOffers_ExchangeOffers"
    elif(recent=="4" and spent=="4" and count=="4"):
         recommend.value  = "Accessories"

    if(warranty=='0'):
            recommend.value = str(recommend.value) +  "_WarrantyRenewOffers"


wb.save('rfm_set.xlsx')


            




    
